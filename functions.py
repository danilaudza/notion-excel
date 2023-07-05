from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import range_boundaries
from variables import currency_style, sum_headers_color, columns_currency

def make_headers(cellRange,worksheet):
    for key in cellRange:
        worksheet.merge_cells(key)
        min_col, min_row, max_col, max_row = range_boundaries(key)

        border = Border(left=Side(border_style='medium'),
                    right=Side(border_style='medium'),
                    top=Side(border_style='medium'),
                    bottom=Side(border_style='medium'))

        cell = worksheet[key[:2]]
        cell.value = cellRange[key]
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4285F4", end_color="4285F4", fill_type="solid")

        for row in worksheet.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.border = border

def convert_currencies(worksheet):
    for column in columns_currency:
        for cell in worksheet[column][2:]:
            cell.style = 'Currency'
            cell.number_format = currency_style

def color_sum_headers(sum_df,worksheet):
    for col_idx, column in enumerate(sum_df.columns):
        cell = worksheet.cell(row=2, column=col_idx + 11)
        cell.fill = PatternFill(start_color=sum_headers_color[col_idx], end_color=sum_headers_color[col_idx], fill_type='solid')
        cell.font = Font(color="FFFFFF")